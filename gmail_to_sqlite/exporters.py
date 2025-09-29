"""
Advanced data export system for Gmail to SQLite.

Provides multiple export formats (CSV, JSON, XML, XLSX, MBOX) with
filtering, transformation, and batch processing capabilities.
"""

import csv
import json
import sqlite3
import xml.etree.ElementTree as ET
from abc import ABC, abstractmethod
from typing import Any, Dict, List, Optional, Union, Iterator
from pathlib import Path
from datetime import datetime
import logging

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportFormat:
    """Export format constants."""
    CSV = "csv"
    JSON = "json"
    XML = "xml"
    XLSX = "xlsx"
    MBOX = "mbox"
    HTML = "html"


class BaseExporter(ABC):
    """Abstract base class for data exporters."""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
    
    @abstractmethod
    def export(self, output_path: str, query: str, params: List[Any] = None, 
               options: Dict[str, Any] = None) -> bool:
        """Export data to specified path."""
        pass
    
    @abstractmethod
    def get_supported_options(self) -> Dict[str, Any]:
        """Get supported export options."""
        pass
    
    def _execute_query(self, query: str, params: List[Any] = None) -> Iterator[sqlite3.Row]:
        """Execute query and yield results."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute(query, params or [])
                
                while True:
                    rows = cursor.fetchmany(1000)  # Batch processing
                    if not rows:
                        break
                    for row in rows:
                        yield row
                        
        except Exception as e:
            logger.error(f"Query execution failed: {e}")
            raise


class CSVExporter(BaseExporter):
    """CSV format exporter."""
    
    def export(self, output_path: str, query: str, params: List[Any] = None, 
               options: Dict[str, Any] = None) -> bool:
        """Export data to CSV format."""
        options = options or {}
        
        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = None
                headers_written = False
                
                for row in self._execute_query(query, params):
                    if not headers_written:
                        fieldnames = list(row.keys())
                        writer = csv.DictWriter(
                            csvfile, 
                            fieldnames=fieldnames,
                            delimiter=options.get('delimiter', ','),
                            quotechar=options.get('quotechar', '"'),
                            quoting=csv.QUOTE_MINIMAL
                        )
                        writer.writeheader()
                        headers_written = True
                    
                    # Convert row to dict and handle JSON fields
                    row_dict = dict(row)
                    for key, value in row_dict.items():
                        if isinstance(value, (dict, list)):
                            row_dict[key] = json.dumps(value)
                        elif value is None:
                            row_dict[key] = options.get('null_value', '')
                    
                    writer.writerow(row_dict)
                
                logger.info(f"CSV export completed: {output_path}")
                return True
                
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return False
    
    def get_supported_options(self) -> Dict[str, Any]:
        """Get supported CSV export options."""
        return {
            'delimiter': ',',
            'quotechar': '"',
            'null_value': ''
        }


class JSONExporter(BaseExporter):
    """JSON format exporter."""
    
    def export(self, output_path: str, query: str, params: List[Any] = None, 
               options: Dict[str, Any] = None) -> bool:
        """Export data to JSON format."""
        options = options or {}
        
        try:
            results = []
            
            for row in self._execute_query(query, params):
                row_dict = dict(row)
                
                # Parse JSON fields
                for key, value in row_dict.items():
                    if isinstance(value, str) and key in ['sender', 'recipients', 'labels']:
                        try:
                            row_dict[key] = json.loads(value)
                        except (json.JSONDecodeError, TypeError):
                            pass  # Keep as string if not valid JSON
                
                results.append(row_dict)
            
            # Write to file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(
                    results, 
                    f, 
                    indent=options.get('indent', 2),
                    ensure_ascii=options.get('ensure_ascii', False),
                    default=str  # Handle datetime objects
                )
            
            logger.info(f"JSON export completed: {output_path} ({len(results)} records)")
            return True
            
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            return False
    
    def get_supported_options(self) -> Dict[str, Any]:
        """Get supported JSON export options."""
        return {
            'indent': 2,
            'ensure_ascii': False
        }


class XMLExporter(BaseExporter):
    """XML format exporter."""
    
    def export(self, output_path: str, query: str, params: List[Any] = None, 
               options: Dict[str, Any] = None) -> bool:
        """Export data to XML format."""
        options = options or {}
        
        try:
            root_name = options.get('root_element', 'messages')
            item_name = options.get('item_element', 'message')
            
            root = ET.Element(root_name)
            root.set('exported_at', datetime.now().isoformat())
            
            for row in self._execute_query(query, params):
                message_elem = ET.SubElement(root, item_name)
                
                for key, value in dict(row).items():
                    if value is not None:
                        elem = ET.SubElement(message_elem, self._sanitize_xml_name(key))
                        
                        if isinstance(value, (dict, list)):
                            # For complex data, store as JSON string
                            elem.text = json.dumps(value)
                            elem.set('type', 'json')
                        else:
                            elem.text = str(value)
            
            # Write XML to file
            tree = ET.ElementTree(root)
            ET.indent(tree, space="  ", level=0)  # Pretty print (Python 3.9+)
            tree.write(output_path, encoding='utf-8', xml_declaration=True)
            
            logger.info(f"XML export completed: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"XML export failed: {e}")
            return False
    
    def _sanitize_xml_name(self, name: str) -> str:
        """Sanitize field name for XML element."""
        # Replace invalid XML name characters
        sanitized = ''.join(c if c.isalnum() or c in '-_.' else '_' for c in name)
        
        # Ensure it starts with a letter or underscore
        if sanitized and not (sanitized[0].isalpha() or sanitized[0] == '_'):
            sanitized = f"field_{sanitized}"
        
        return sanitized or 'unknown_field'
    
    def get_supported_options(self) -> Dict[str, Any]:
        """Get supported XML export options."""
        return {
            'root_element': 'messages',
            'item_element': 'message'
        }


class XLSXExporter(BaseExporter):
    """Excel XLSX format exporter."""
    
    def export(self, output_path: str, query: str, params: List[Any] = None, 
               options: Dict[str, Any] = None) -> bool:
        """Export data to XLSX format."""
        if not XLSX_AVAILABLE:
            logger.error("openpyxl not available for XLSX export")
            return False
        
        options = options or {}
        
        try:
            # Use pandas for easier Excel export if available
            if PANDAS_AVAILABLE:
                return self._export_with_pandas(output_path, query, params, options)
            else:
                return self._export_with_openpyxl(output_path, query, params, options)
                
        except Exception as e:
            logger.error(f"XLSX export failed: {e}")
            return False
    
    def _export_with_pandas(self, output_path: str, query: str, 
                           params: List[Any], options: Dict[str, Any]) -> bool:
        """Export using pandas (more efficient for large datasets)."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                df = pd.read_sql_query(query, conn, params=params)
                
                # Handle JSON columns
                json_columns = ['sender', 'recipients', 'labels']
                for col in json_columns:
                    if col in df.columns:
                        df[col] = df[col].apply(
                            lambda x: json.dumps(x) if isinstance(x, (dict, list)) else x
                        )
                
                # Write to Excel
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(
                        writer, 
                        sheet_name=options.get('sheet_name', 'Messages'),
                        index=options.get('include_index', False)
                    )
                
                logger.info(f"XLSX export completed: {output_path} ({len(df)} records)")
                return True
                
        except Exception as e:
            logger.error(f"Pandas XLSX export failed: {e}")
            return False
    
    def _export_with_openpyxl(self, output_path: str, query: str,
                             params: List[Any], options: Dict[str, Any]) -> bool:
        """Export using openpyxl directly."""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = options.get('sheet_name', 'Messages')
        
        headers_written = False
        row_num = 1
        
        for row in self._execute_query(query, params):
            if not headers_written:
                # Write headers
                for col_num, header in enumerate(row.keys(), 1):
                    ws.cell(row=row_num, column=col_num, value=header)
                row_num += 1
                headers_written = True
            
            # Write data
            for col_num, value in enumerate(dict(row).values(), 1):
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_num, column=col_num, value=value)
            
            row_num += 1
        
        wb.save(output_path)
        logger.info(f"XLSX export completed: {output_path}")
        return True
    
    def get_supported_options(self) -> Dict[str, Any]:
        """Get supported XLSX export options."""
        return {
            'sheet_name': 'Messages',
            'include_index': False
        }


class MBOXExporter(BaseExporter):
    """MBOX format exporter (standard email archive format)."""
    
    def export(self, output_path: str, query: str, params: List[Any] = None, 
               options: Dict[str, Any] = None) -> bool:
        """Export data to MBOX format."""
        options = options or {}
        
        try:
            with open(output_path, 'w', encoding='utf-8') as mbox_file:
                for row in self._execute_query(query, params):
                    message_data = dict(row)
                    mbox_message = self._format_as_mbox_message(message_data, options)
                    mbox_file.write(mbox_message)
                    mbox_file.write('\n')
            
            logger.info(f"MBOX export completed: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"MBOX export failed: {e}")
            return False
    
    def _format_as_mbox_message(self, message_data: Dict[str, Any], 
                               options: Dict[str, Any]) -> str:
        """Format message data as MBOX message."""
        # MBOX format starts with "From " line
        sender = self._extract_sender_email(message_data.get('sender'))
        timestamp = message_data.get('timestamp', datetime.now().isoformat())
        
        # Convert timestamp to MBOX format
        try:
            dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
            mbox_date = dt.strftime('%a %b %d %H:%M:%S %Y')
        except:
            mbox_date = datetime.now().strftime('%a %b %d %H:%M:%S %Y')
        
        lines = [f"From {sender} {mbox_date}"]
        
        # Add headers
        lines.append(f"Message-ID: <{message_data.get('message_id', 'unknown')}>")
        lines.append(f"Date: {timestamp}")
        lines.append(f"From: {sender}")
        
        # Add recipients
        recipients = message_data.get('recipients', {})
        if isinstance(recipients, str):
            try:
                recipients = json.loads(recipients)
            except:
                recipients = {}
        
        if isinstance(recipients, dict):
            for header, recipient_list in recipients.items():
                if recipient_list and header.lower() in ['to', 'cc', 'bcc']:
                    recipient_emails = []
                    for recipient in recipient_list:
                        if isinstance(recipient, dict):
                            email = recipient.get('email', '')
                            name = recipient.get('name', '')
                            if name:
                                recipient_emails.append(f'"{name}" <{email}>')
                            else:
                                recipient_emails.append(email)
                        else:
                            recipient_emails.append(str(recipient))
                    
                    if recipient_emails:
                        lines.append(f"{header.title()}: {', '.join(recipient_emails)}")
        
        # Add subject
        subject = message_data.get('subject', '(No Subject)')
        lines.append(f"Subject: {subject}")
        
        # Add other headers
        lines.append("Content-Type: text/plain; charset=utf-8")
        lines.append("")  # Empty line separates headers from body
        
        # Add body
        body = message_data.get('body', '')
        if body:
            # Escape "From " lines in body (MBOX requirement)
            body_lines = body.split('\n')
            escaped_lines = []
            for line in body_lines:
                if line.startswith('From '):
                    escaped_lines.append('>' + line)
                else:
                    escaped_lines.append(line)
            lines.extend(escaped_lines)
        
        return '\n'.join(lines)
    
    def _extract_sender_email(self, sender_data: Any) -> str:
        """Extract email address from sender data."""
        if isinstance(sender_data, str):
            try:
                sender_dict = json.loads(sender_data)
                return sender_dict.get('email', 'unknown@example.com')
            except:
                return sender_data if '@' in sender_data else 'unknown@example.com'
        elif isinstance(sender_data, dict):
            return sender_data.get('email', 'unknown@example.com')
        else:
            return 'unknown@example.com'
    
    def get_supported_options(self) -> Dict[str, Any]:
        """Get supported MBOX export options."""
        return {}


class HTMLExporter(BaseExporter):
    """HTML format exporter with styling."""
    
    def export(self, output_path: str, query: str, params: List[Any] = None, 
               options: Dict[str, Any] = None) -> bool:
        """Export data to HTML format."""
        options = options or {}
        
        try:
            html_content = self._generate_html_report(query, params, options)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"HTML export completed: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"HTML export failed: {e}")
            return False
    
    def _generate_html_report(self, query: str, params: List[Any], 
                             options: Dict[str, Any]) -> str:
        """Generate HTML report content."""
        title = options.get('title', 'Gmail Messages Report')
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background: #f4f4f4; padding: 15px; margin-bottom: 20px; }}
                .message {{ border: 1px solid #ddd; margin: 10px 0; padding: 15px; }}
                .message-header {{ background: #e8f4f8; padding: 10px; margin: -15px -15px 15px -15px; }}
                .message-body {{ white-space: pre-wrap; }}
                .metadata {{ font-size: 0.9em; color: #666; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        """
        
        # Add messages
        message_count = 0
        for row in self._execute_query(query, params):
            message_data = dict(row)
            html += self._format_message_html(message_data)
            message_count += 1
        
        if message_count == 0:
            html += "<p>No messages found matching the criteria.</p>"
        else:
            html += f"<p class='metadata'>Total messages: {message_count}</p>"
        
        html += """
        </body>
        </html>
        """
        
        return html
    
    def _format_message_html(self, message_data: Dict[str, Any]) -> str:
        """Format single message as HTML."""
        sender = message_data.get('sender', {})
        if isinstance(sender, str):
            try:
                sender = json.loads(sender)
            except:
                sender = {'email': sender}
        
        sender_display = sender.get('name', sender.get('email', 'Unknown'))
        
        timestamp = message_data.get('timestamp', '')
        subject = message_data.get('subject', '(No Subject)')
        body = message_data.get('body', '')
        
        # Truncate long bodies for display
        if len(body) > 1000:
            body = body[:1000] + "... [truncated]"
        
        return f"""
        <div class="message">
            <div class="message-header">
                <strong>{subject}</strong><br>
                <span class="metadata">From: {sender_display} | Date: {timestamp}</span>
            </div>
            <div class="message-body">{body}</div>
        </div>
        """
    
    def get_supported_options(self) -> Dict[str, Any]:
        """Get supported HTML export options."""
        return {
            'title': 'Gmail Messages Report'
        }


class ExportManager:
    """Main export manager with support for multiple formats."""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.exporters = {
            ExportFormat.CSV: CSVExporter(db_path),
            ExportFormat.JSON: JSONExporter(db_path),
            ExportFormat.XML: XMLExporter(db_path),
            ExportFormat.XLSX: XLSXExporter(db_path),
            ExportFormat.MBOX: MBOXExporter(db_path),
            ExportFormat.HTML: HTMLExporter(db_path)
        }
    
    def export(self, format_type: str, output_path: str, 
               query: Optional[str] = None, params: List[Any] = None,
               options: Dict[str, Any] = None) -> bool:
        """Export data in specified format."""
        if format_type not in self.exporters:
            logger.error(f"Unsupported export format: {format_type}")
            return False
        
        # Use default query if none provided
        if query is None:
            query = "SELECT * FROM messages WHERE is_deleted = 0 ORDER BY timestamp DESC"
        
        exporter = self.exporters[format_type]
        return exporter.export(output_path, query, params or [], options or {})
    
    def export_filtered_messages(self, format_type: str, output_path: str,
                                filters: Dict[str, Any] = None,
                                options: Dict[str, Any] = None) -> bool:
        """Export messages with filters applied."""
        where_clauses = []
        params = []
        
        # Build WHERE clause from filters
        if filters:
            for field, value in filters.items():
                if field == 'date_range' and isinstance(value, tuple):
                    where_clauses.append("timestamp BETWEEN ? AND ?")
                    params.extend(value)
                elif field == 'sender_email':
                    where_clauses.append("json_extract(sender, '$.email') LIKE ?")
                    params.append(f"%{value}%")
                elif field == 'subject_contains':
                    where_clauses.append("subject LIKE ?")
                    params.append(f"%{value}%")
                elif field == 'has_label':
                    where_clauses.append("json_extract(labels, '$') LIKE ?")
                    params.append(f"%{value}%")
                elif field == 'is_unread':
                    where_clauses.append("is_read = ?")
                    params.append(0 if value else 1)
                elif field == 'min_size_mb':
                    min_bytes = int(value * 1024 * 1024)
                    where_clauses.append("size >= ?")
                    params.append(min_bytes)
        
        # Build complete query
        base_query = "SELECT * FROM messages WHERE is_deleted = 0"
        if where_clauses:
            base_query += " AND " + " AND ".join(where_clauses)
        base_query += " ORDER BY timestamp DESC"
        
        return self.export(format_type, output_path, base_query, params, options)
    
    def get_available_formats(self) -> List[str]:
        """Get list of available export formats."""
        return list(self.exporters.keys())
    
    def get_format_options(self, format_type: str) -> Dict[str, Any]:
        """Get available options for a specific format."""
        if format_type in self.exporters:
            return self.exporters[format_type].get_supported_options()
        return {}
    
    def export_contact_summary(self, format_type: str, output_path: str,
                              options: Dict[str, Any] = None) -> bool:
        """Export contact summary report."""
        query = """
        SELECT 
            json_extract(sender, '$.email') as email,
            json_extract(sender, '$.name') as name,
            COUNT(*) as total_messages,
            SUM(CASE WHEN is_outgoing = 1 THEN 1 ELSE 0 END) as sent_count,
            SUM(CASE WHEN is_outgoing = 0 THEN 1 ELSE 0 END) as received_count,
            MAX(timestamp) as last_contact,
            SUM(size) as total_size,
            AVG(size) as avg_size
        FROM messages 
        WHERE is_deleted = 0 
            AND json_extract(sender, '$.email') IS NOT NULL
            AND json_extract(sender, '$.email') != ''
        GROUP BY json_extract(sender, '$.email')
        ORDER BY total_messages DESC
        LIMIT 100
        """
        
        return self.export(format_type, output_path, query, [], options)